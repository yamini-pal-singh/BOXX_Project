"""
Pytest configuration and fixtures for the BOXX Chatbot Test Automation Suite.
"""

import csv
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger = logging.getLogger("conftest")
    logger.warning("openpyxl not installed — Excel output will be skipped. Run: pip install openpyxl")

import pytest

from boxx_client import BOXXClient
from test_loader import load_test_cases, load_sample_cases

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stderr,
)

logger = logging.getLogger("conftest")

# Session-level tracking
_session_start: float = 0.0


def _get_results_dir() -> Path:
    """Return the results/ directory, creating it if needed."""
    results_dir = Path(__file__).parent / "results"
    results_dir.mkdir(parents=True, exist_ok=True)
    return results_dir


# ---------------------------------------------------------------------------
# pytest hooks
# ---------------------------------------------------------------------------


def pytest_addoption(parser):
    parser.addoption(
        "--test-sheet",
        action="store",
        default=None,
        help="Path to the test case sheet (CSV or XLSX). Uses built-in samples if omitted.",
    )
    parser.addoption(
        "--base-url",
        action="store",
        default=None,
        help="BOXX API base URL (overrides env/ default).",
    )
    parser.addoption(
        "--api-key",
        action="store",
        default=None,
        help="BOXX API key (overrides env/ default).",
    )
    parser.addoption(
        "--results-dir",
        action="store",
        default=None,
        help="Custom directory for results CSV output (default: ./results/)",
    )


def pytest_configure(config):
    """Register custom markers and configure."""
    global _session_start
    _session_start = __import__("time").time()
    config.addinivalue_line("markers", "smoke: quick smoke test to verify API reachability.")
    config.addinivalue_line("markers", "sample: test using built-in sample data (runs without a sheet file).")
    # Store start time in config for hooks
    config._boxx_start = _session_start  # type: ignore[attr-defined]


def pytest_collectstart():
    """Ignore non-test classes that pytest may try to collect."""
    pass  # handled via __test__ attribute below


# Prevent TestCase dataclass from being collected as a test
import test_loader as _tl
_tl.TestCase.__test__ = False  # type: ignore[attr-defined]


def pytest_generate_tests(metafunc):
    """Parametrize 'test_case' fixtures from loaded test data at collection time."""
    if "test_case" not in metafunc.fixturenames:
        return

    sheet_path = metafunc.config.getoption("--test-sheet", default=None)
    if sheet_path:
        path = Path(sheet_path)
        if not path.exists():
            # Try relative to project root
            alt = Path(__file__).parent / sheet_path
            if alt.exists():
                path = alt
            else:
                pytest.exit(f"Test sheet not found: {sheet_path}", returncode=1)
        cases = load_test_cases(path)
    else:
        logger.info("No --test-sheet provided; using built-in sample test cases.")
        cases = load_sample_cases()

    if not cases:
        pytest.exit("No test cases loaded!", returncode=1)

    metafunc.parametrize(
        "test_case",
        cases,
        ids=[c.test_id for c in cases],
    )


def _write_excel_results(
    filepath: Path,
    test_results: list[dict],
    fieldnames: list[str],
    total: int,
    passed: int,
    failed: int,
    skipped: int,
    duration_str: str,
    source_desc: str,
):
    """Write results to a properly formatted Excel workbook.

    Creates a Results sheet (data table with color-coding) and a Summary sheet.
    """
    if not HAS_OPENPYXL:
        logger.warning("openpyxl not available — skipping Excel output.")
        return

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pass_font = Font(color="006100")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_font = Font(color="9C0006")
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    skip_font = Font(color="9C6500")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Determine column widths based on field names
    col_widths = {
        "timestamp": 22, "test_id": 18, "language": 10, "scenario_type": 14,
        "status": 10, "turn_count": 10, "input_messages": 50,
        "expected_keywords": 30, "expected_classification": 22, "expected_journey": 16,
        "expected_emotion": 16, "latency_ms": 10, "classification": 22, "journey": 16,
        "emotion": 16, "emotion_language": 16, "loss_assessment": 20,
        "persona_archetype": 22, "persona_confidence": 18, "response_style": 16,
        "keywords_matched": 30, "expected_not_found": 30,
        "bot_reply": 60, "full_analysis": 60,
        "error_message": 50, "notes": 30, "source_sheet": 22,
        "session_id": 28, "reply_count": 12,
        "final_reply": 60, "final_classification": 22, "final_journey": 16, "final_emotion": 16,
        "conversation_log": 80,
    }

    # ── Sheet 1: Results ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Results"

    # Header row
    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, result in enumerate(test_results, start=2):
        for col_idx, name in enumerate(fieldnames, start=1):
            value = result.get(name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx > 7))

            # Color by status column
            status = str(result.get("status", "")).strip().upper()
            if status == "PASS":
                cell.fill = pass_fill
                cell.font = pass_font
            elif status == "FAIL":
                cell.fill = fail_fill
                cell.font = fail_font
            elif status == "SKIP":
                cell.fill = skip_fill
                cell.font = skip_font

    # Column widths
    for col_idx, name in enumerate(fieldnames, start=1):
        width = col_widths.get(name, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(test_results) + 1}"

    # ── Sheet 2: Summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    label_font = Font(name="Calibri", bold=True, size=11)
    value_font = Font(name="Calibri", size=11)
    stats_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws2.cell(row=1, column=1, value="BOXX Test Automation — Results Summary").font = title_font
    ws2.merge_cells("A1:B1")

    stats = [
        ("Run Timestamp", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Test Source", source_desc),
        ("Duration", duration_str),
        ("", ""),
        ("Total Tests", str(total)),
        ("Passed", str(passed)),
        ("Failed", str(failed)),
        ("Skipped", str(skipped)),
    ]

    for i, (label, value) in enumerate(stats, start=3):
        cell_l = ws2.cell(row=i, column=1, value=label)
        cell_l.font = label_font
        cell_l.fill = stats_fill
        cell_v = ws2.cell(row=i, column=2, value=value)
        cell_v.font = value_font

        # Color the pass/fail values
        if label == "Passed":
            cell_v.font = Font(name="Calibri", bold=True, size=12, color="006100")
            cell_v.fill = pass_fill
        elif label == "Failed":
            cell_v.font = Font(name="Calibri", bold=True, size=12, color="9C0006")
            cell_v.fill = fail_fill
        elif label == "Skipped":
            cell_v.font = Font(name="Calibri", bold=True, size=12, color="9C6500")
            cell_v.fill = skip_fill

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 40

    # Save
    wb.save(str(filepath))
    logger.info("Excel results also written to: %s", filepath)


def pytest_sessionfinish(session, exitstatus):
    """Write all accumulated test results to a timestamped CSV."""
    # Import results from test module
    from test_boxx_scam_flows import TEST_RESULTS

    if not TEST_RESULTS:
        logger.info("No test results to write.")
        return

    results_dir = _get_results_dir()
    config = session.config
    custom_dir = config.getoption("--results-dir", default=None)
    if custom_dir:
        results_dir = Path(custom_dir)
        results_dir.mkdir(parents=True, exist_ok=True)

    # Compute session duration
    start = getattr(config, "_boxx_start", 0)
    duration_s = __import__("time").time() - start
    duration_str = f"{duration_s / 60:.1f} min" if duration_s > 120 else f"{duration_s:.0f}s"

    # Count stats
    total = len(TEST_RESULTS)
    passed = sum(1 for r in TEST_RESULTS if r["status"] == "PASS")
    failed = sum(1 for r in TEST_RESULTS if r["status"] in ("FAIL", "ERROR"))
    skipped = sum(1 for r in TEST_RESULTS if r["status"] == "SKIP")

    # Determine test source
    sheet_path = config.getoption("--test-sheet", default=None)
    source_desc = sheet_path or "built-in samples"

    # Timestamp for filename
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"boxx_results_{ts}.csv"
    filepath = results_dir / filename

    # Collect all field names from results (in a consistent order)
    fieldnames = [
        "timestamp", "test_id", "language", "scenario_type",
        "status", "turn_count", "input_messages",
        "expected_keywords", "expected_classification", "expected_journey", "expected_emotion",
        "latency_ms", "classification", "journey", "emotion", "emotion_language",
        "loss_assessment", "persona_archetype", "persona_confidence", "response_style",
        "keywords_matched", "expected_not_found",
        "bot_reply", "full_analysis",
        "error_message", "notes", "source_sheet",
        # New fields
        "session_id", "reply_count",
        "final_reply", "final_classification", "final_journey", "final_emotion",
        "conversation_log",
    ]

    # Write CSV
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")

        # ── Metadata header ──
        f.write(f"# BOXX Test Automation — Results\n")
        f.write(f"# Run Timestamp: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}\n")
        f.write(f"# Test Source:   {source_desc}\n")
        f.write(f"# Duration:      {duration_str}\n")
        f.write(f"# Total Tests:   {total}\n")
        f.write(f"# Passed:        {passed}\n")
        f.write(f"# Failed:        {failed}\n")
        f.write(f"# Skipped:       {skipped}\n")
        f.write(f"# API Base URL:  {os.environ.get('BOXX_BASE_URL', 'https://boxxv2.shunyalabs.ai')}\n")
        f.write(f"#\n")

        writer.writeheader()
        for row in TEST_RESULTS:
            writer.writerow(row)

        # ── Summary footer ──
        f.write(f"\n# SUMMARY: TOTAL={total}  PASS={passed}  FAIL={failed}  SKIP={skipped}  DURATION={duration_str}\n")

    logger.info("\n"
                "╔══════════════════════════════════════════════════════════╗\n"
                "║  Results written to: %-32s  ║\n"
                "║  TOTAL: %-4d  |  PASS: %-4d  |  FAIL: %-4d  |  SKIP: %-4d       ║\n"
                "║  Duration: %-65s ║\n"
                "╚══════════════════════════════════════════════════════════╝"
                "", filepath, total, passed, failed, skipped, duration_str)

    # Also write a latest symlink/copy (overwrite the "latest" file)
    latest_path = results_dir / "boxx_results_latest.csv"
    import shutil
    shutil.copy2(filepath, latest_path)
    logger.info("Latest results also available at: %s", latest_path)

    # ── Write Excel results ──────────────────────────────────────────
    if HAS_OPENPYXL:
        excel_filename = f"boxx_results_{ts}.xlsx"
        excel_path = results_dir / excel_filename
        _write_excel_results(
            filepath=excel_path,
            test_results=TEST_RESULTS,
            fieldnames=fieldnames,
            total=total,
            passed=passed,
            failed=failed,
            skipped=skipped,
            duration_str=duration_str,
            source_desc=source_desc,
        )
        # Also copy to latest Excel
        latest_excel = results_dir / "boxx_results_latest.xlsx"
        shutil.copy2(excel_path, latest_excel)
    else:
        logger.warning("Skipping Excel output — install openpyxl: pip install openpyxl")


# ---------------------------------------------------------------------------
# Session-scoped fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def boxx_client(request) -> BOXXClient:
    """Create a single BOXXClient for the whole session."""
    base_url = request.config.getoption("--base-url") or None
    api_key = request.config.getoption("--api-key") or None
    client = BOXXClient(base_url=base_url, api_key=api_key)
    # Run health check once at session start
    ok = client.health_check()
    if not ok:
        pytest.exit(
            "BOXX API health check FAILED.\n"
            f"  URL: {client.base_url}\n"
            "  Check that the API is running and your API key is valid.\n"
            "  Set BOXX_BASE_URL / BOXX_API_KEY env vars if defaults are wrong.",
            returncode=1,
        )
    return client
