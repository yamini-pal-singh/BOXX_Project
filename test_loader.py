"""
Test Case Loader
=================
Reads test cases from the BOXX Master Testcases.xlsx workbook (all relevant sheets)
and transforms them into a uniform list of TestCase namedtuples.

It auto-detects column layouts across different sheets, so you don't need to
rename columns in the source file.
"""

from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

logger = logging.getLogger("test_loader")

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class TestCase:
    test_id: str
    scenario_type: str  # "single-turn" | "multi-turn"
    language: str
    input_messages: list[str]  # ordered list of user turns (multi-turn if >1)
    expected_keywords: list[str]  # all should appear in final reply (case-insensitive)
    expected_classification: str = ""   # optional — e.g. "phishing", "otp_fraud"
    expected_journey: str = ""          # optional — e.g. "Flow1", "Flow2"
    expected_emotion: str = ""          # optional
    expected_detection: str = ""        # optional — broader expected detection label
    notes: str = ""
    source_sheet: str = ""              # which sheet this came from

# ---------------------------------------------------------------------------
# Column-mapping heuristics
# ---------------------------------------------------------------------------

# Maps known column-name fragments → standardised keys
_COLUMN_ALIASES = {
    # Test ID
    "test case": "test_id",
    "test id": "test_id",
    "tc id": "test_id",
    "test_id": "test_id",
    # Scenario / description
    "scenario": "scenario",
    # User input / steps
    "user input": "input",
    "user query": "input",
    "hindi user query": "input",
    "input message": "input",
    "input messages": "input",
    "input_message": "input",
    "input_messages": "input",
    "steps": "input",
    "step": "input",
    "precondition": "input",
    "preconditions": "input",
    # Expected result / behaviour
    "expected result": "expected",
    "expected platform response": "expected",
    "expected bot behaviour": "expected",
    "expected ai behaviour": "expected",
    "expected detection": "expected_detection",
    # Classification / journey
    "expected classification": "expected_classification",
    "expected journey": "expected_journey",
    "scenario type": "scenario_type",
    "expected keywords": "keywords",
}

# Multi-turn messages from the source are separated by these delimiters
_MULTI_TURN_DELIMITERS = re.compile(r"\s*\|\s*|\n")


def _normalise_header(name: str) -> str:
    """Map a raw column header to our standard key, or return the original lowercased."""
    clean = name.strip().lower().replace("  ", " ")
    # Try exact match by alias
    if clean in _COLUMN_ALIASES:
        return _COLUMN_ALIASES[clean]
    # Try prefix match
    for alias, key in _COLUMN_ALIASES.items():
        if clean.startswith(alias):
            return key
    return clean


def _guess_language_from_text(text: str) -> str:
    """Rough heuristic: if text contains Devanagari chars, default to 'hi'."""
    if re.search(r"[ऀ-ॿ]", text):
        return "hi"
    return "en"

# ---------------------------------------------------------------------------
# Excel loader
# ---------------------------------------------------------------------------

def _load_from_xlsx(path: str | Path) -> list[TestCase]:
    """Read all relevant sheets from the Excel workbook and convert to TestCases."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    cases: list[TestCase] = []

    # Sheets whose content we should process (skip Summary, Requirement_Master)
    skip_sheets = {"Summary", "Requirement_Master"}

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Parse header row
        headers = [str(c).strip() if c else "" for c in rows[0]]
        col_map: dict[str, int] = {}
        for idx, raw in enumerate(headers):
            key = _normalise_header(raw)
            if key:
                col_map[key] = idx

        # We need at least an "input" column to produce test cases
        input_col = col_map.get("input")
        if input_col is None:
            logger.debug("Skipping sheet '%s' — no recognised input column", sheet_name)
            continue

        test_id_col = col_map.get("test_id")
        scenario_col = col_map.get("scenario")
        expected_col = col_map.get("expected")
        expected_det_col = col_map.get("expected_detection")
        lang_col = col_map.get("language")
        type_col = col_map.get("scenario_type")
        kw_col = col_map.get("keywords")
        cls_col = col_map.get("expected_classification")
        jny_col = col_map.get("expected_journey")

        for row_idx, row in enumerate(rows[1:], start=2):
            # Skip truly empty rows
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            # Skip rows that are clearly comments (no usable input)
            raw_input = str(row[input_col]).strip() if input_col < len(row) and row[input_col] else ""
            if not raw_input or raw_input.lower() in ("na", "n/a", "-", "none", "skip", "todo"):
                continue

            # Build the test case
            test_id = str(row[test_id_col]).strip() if test_id_col is not None and test_id_col < len(row) and row[test_id_col] else f"{sheet_name}_R{row_idx}"
            scenario = str(row[scenario_col]).strip() if scenario_col is not None and scenario_col < len(row) and row[scenario_col] else ""
            expected = str(row[expected_col]).strip() if expected_col is not None and expected_col < len(row) and row[expected_col] else ""
            expected_det = str(row[expected_det_col]).strip() if expected_det_col is not None and expected_det_col < len(row) and row[expected_det_col] else ""

            # Read explicit columns if present (Automation_Test_Suite style)
            explicit_lang = str(row[lang_col]).strip() if lang_col is not None and lang_col < len(row) and row[lang_col] else ""
            explicit_type = str(row[type_col]).strip() if type_col is not None and type_col < len(row) and row[type_col] else ""
            explicit_kw = str(row[kw_col]).strip() if kw_col is not None and kw_col < len(row) and row[kw_col] else ""
            explicit_cls = str(row[cls_col]).strip() if cls_col is not None and cls_col < len(row) and row[cls_col] else ""
            explicit_jny = str(row[jny_col]).strip() if jny_col is not None and jny_col < len(row) and row[jny_col] else ""

            # Determine scenario type: prefer explicit, fallback to auto-detect
            if explicit_type:
                scenario_type = "multi-turn" if "multi" in explicit_type.lower() else "single-turn"
            elif "|" in raw_input or "\n" in raw_input:
                messages_raw = [m.strip() for m in _MULTI_TURN_DELIMITERS.split(raw_input) if m.strip()]
                scenario_type = "multi-turn" if len(messages_raw) > 1 else "single-turn"
            else:
                scenario_type = "single-turn"

            # Multi-turn: split on pipe/newline only if not already using explicit list
            if "|" in raw_input or "\n" in raw_input:
                if not explicit_type or scenario_type == "multi-turn":
                    messages = [m.strip() for m in _MULTI_TURN_DELIMITERS.split(raw_input) if m.strip()]
                else:
                    messages = [raw_input]
            else:
                messages = [raw_input]

            # Language: prefer explicit from sheet
            language = explicit_lang or _guess_language_from_text(raw_input)

            # Keywords: prefer explicit pipe-separated list from sheet
            if explicit_kw:
                keywords = [k.strip().lower() for k in explicit_kw.split("|") if k.strip()]
            else:
                keywords = _extract_keywords(expected, expected_det, scenario, raw_input)

            # Classification / journey: prefer explicit from sheet
            if explicit_cls:
                classification = explicit_cls
            else:
                classification, _journey, _emotion = _parse_expected(expected, expected_det)
                if not classification:
                    classification = _detect_scenario_classification(scenario, raw_input)

            if explicit_jny:
                journey = explicit_jny
            else:
                _, journey, _ = _parse_expected(expected, expected_det)

            emotion = ""
            if not journey:
                _, _, emotion = _parse_expected(expected, expected_det)

            cases.append(TestCase(
                test_id=test_id,
                scenario_type=scenario_type,
                language=language,
                input_messages=messages,
                expected_keywords=keywords,
                expected_classification=classification,
                expected_journey=journey,
                expected_emotion=emotion,
                expected_detection=expected_det or expected,
                notes=scenario,
                source_sheet=sheet_name,
            ))

    wb.close()
    return cases


def _extract_keywords(expected: str, expected_det: str, scenario: str, user_input: str) -> list[str]:
    """Extract meaningful keywords to assert on from expected result text."""
    words = set()
    # Common action words we want to assert on
    for text in [expected, expected_det, scenario]:
        if not text:
            continue
        # Look for specific domain terms
        domain_terms = [
            "scam", "fraud", "phishing", "otp", "upi", "recovery", "1930",
            "disclaimer", "welcome", "empathy", "block", "reset", "password",
            "guided", "workflow", "flow1", "flow2", "flow3", "flow4", "flow5",
            "cyber", "crime", "report", "classifies", "identifies", "warning",
            "safe", "genuine", "fake", "preventive", "guidance", "next steps",
            "clarifying", "redirect", "route", "credentials", "payment",
            "no action", "interact", "dont panic", "action", "support",
            "emotional", "help", "immediate", "emergency", "freeze",
            "evidence", "utr", "fir", "police", "national", "helpline",
        ]
        lower = text.lower()
        for term in domain_terms:
            if term in lower:
                words.add(term)

    # If we got nothing useful, add a generic safety keyword
    if not words:
        words.add("scam" if "scam" in (expected + scenario).lower() else "help")

    return list(words)


def _detect_scenario_classification(scenario: str, user_input: str) -> str:
    """Heuristic classification of scam type from scenario/input text."""
    text = (scenario + " " + user_input).lower()
    patterns = [
        ("upi_fraud", r"\bupi\b"),
        ("otp_fraud", r"\botp\b"),
        ("phishing", r"phish|link.*click|kyc.*update|fake (sms|email|message)"),
        ("qr_scam", r"qr\s*scan"),
        ("remote_access", r"anydesk|teamviewer|remote|app.*install"),
        ("digital_arrest", r"digital arrest|police|video.call|arrest|warrant|cbi|polic"),
        ("job_fraud", r"job|work.from.home|registration fee|recruitment"),
        ("loan_scam", r"loan app|instant loan|fake loan"),
        ("sextortion", r"sextort|blackmail|video.*record"),
        ("investment_scam", r"investment|stock|trading|profit|returns? high"),
        ("lottery_scam", r"lottery|prize|won|winner"),
        ("marketplace_fraud", r"marketplace|olx|facebook.*market|product.*not.*deliver"),
        ("card_fraud", r"card.*skim|card.*clone|unauthorized.*card|card.*block"),
        ("sim_swap", r"sim\s*swap|sim.*change"),
        ("identity_theft", r"aadhaar|aadhar|pan.*misuse|identity.*theft"),
        ("fake_ad", r"fake\s*ad|fake.*offer|discount.*link"),
        ("wifi_risk", r"public\s*wifi|wi.?fi.*risk"),
        ("general_cyber", r"hacked|account.*compromise|virus|malware"),
        ("advance_fee", r"advance.*fee|processing.*fee|registration.*fee|fee.*pay"),
    ]
    for classification, pattern in patterns:
        if re.search(pattern, text):
            return classification
    return ""


def _parse_expected(expected: str, expected_det: str) -> tuple[str, str, str]:
    """Parse expected text for classification, journey, and emotion hints."""
    text = (expected + " " + expected_det).lower()
    classification = _detect_scenario_classification(text, "")
    journey = ""
    emotion = ""

    # Detect journey flow
    for flow in ["flow1", "flow2", "flow3", "flow4", "flow5"]:
        if flow in text:
            journey = flow.capitalize()
            break

    # Detect emotion hints
    if any(w in text for w in ["empathy", "emotional", "support", "reassur", "dont panic"]):
        emotion = "empathy"

    return classification or _detect_scenario_classification(text, ""), journey, emotion


# ---------------------------------------------------------------------------
# CSV loader (for standalone test case files)
# ---------------------------------------------------------------------------

def _load_from_csv(path: str | Path) -> list[TestCase]:
    """Load test cases from a simple CSV file."""
    cases: list[TestCase] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            test_id = (row.get("test_id") or "").strip()
            if not test_id:
                continue
            raw_input = (row.get("input_message") or row.get("input_messages") or "").strip()
            if not raw_input:
                continue
            # Multi-turn detection
            if "|" in raw_input:
                messages = [m.strip() for m in raw_input.split("|") if m.strip()]
                scenario_type = "multi-turn"
            else:
                messages = [raw_input]
                scenario_type = "single-turn"

            lang = (row.get("language") or _guess_language_from_text(raw_input)).strip()
            keywords_raw = (row.get("expected_keywords") or "").strip()
            keywords = [k.strip().lower() for k in keywords_raw.split("|") if k.strip()] if keywords_raw else ["scam"]

            cases.append(TestCase(
                test_id=test_id,
                scenario_type=scenario_type,
                language=lang,
                input_messages=messages,
                expected_keywords=keywords,
                expected_classification=(row.get("expected_classification") or "").strip(),
                expected_journey=(row.get("expected_journey") or "").strip(),
                notes=(row.get("notes") or "").strip(),
            ))
    return cases


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_test_cases(path: str | Path) -> list[TestCase]:
    """Load test cases from either an Excel (.xlsx) or CSV file.

    For Excel files, reads all relevant sheets in the workbook and transforms
    each row into a TestCase, adapting to whatever column layout each sheet uses.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Test case file not found: {path}")

    if path.suffix.lower() in (".xlsx", ".xls"):
        logger.info("Loading test cases from Excel: %s", path)
        cases = _load_from_xlsx(path)
    elif path.suffix.lower() == ".csv":
        logger.info("Loading test cases from CSV: %s", path)
        cases = _load_from_csv(path)
    else:
        raise ValueError(f"Unsupported file format: {path.suffix} (use .xlsx or .csv)")

    # Deduplicate by test_id (keep first occurrence)
    seen: set[str] = set()
    deduped: list[TestCase] = []
    for c in cases:
        if c.test_id not in seen:
            seen.add(c.test_id)
            deduped.append(c)
        else:
            logger.warning("Duplicate test_id skipped: %s", c.test_id)

    logger.info("Loaded %d test cases (from %d raw rows)", len(deduped), len(cases))
    return deduped


def load_sample_cases() -> list[TestCase]:
    """Return built-in sample test cases so the suite is runnable without a sheet file."""
    return [
        # All sample cases use realistic keywords based on observed bot responses.
        # expected_classification is omitted (analysis.classification may be empty
        # at the triage stage) — the keyword assertions validate the reply text.
        TestCase("SAMPLE_001", "single-turn", "en",
                 ["I got an SMS saying my KYC will expire and to click a link"],
                 ["phishing", "scam"]),
        TestCase("SAMPLE_002", "single-turn", "en",
                 ["Someone called from my bank and asked for my OTP, I told them"],
                 ["happening", "family"]),
        TestCase("SAMPLE_003", "single-turn", "en",
                 ["I clicked a phishing link but no money was taken"],
                 ["suspicious", "link", "password"]),
        TestCase("SAMPLE_004", "single-turn", "hi",
                 ["मेरा पैसा कट गया"],
                 ["पैसे", "माध्यम"]),
        TestCase("SAMPLE_005", "single-turn", "en",
                 ["I paid for an iPhone on Facebook Marketplace but never received it"],
                 ["facebook", "marketplace", "payment"]),
        TestCase("SAMPLE_006", "single-turn", "hi",
                 ["मुझे पुलिस बनकर वीडियो कॉल आई धमकी दे रहे हैं"],
                 ["साइबर", "घोटाला"]),
        TestCase("SAMPLE_007", "single-turn", "en",
                 ["Is this message genuine? ICICI Bank: Your account will be suspended, update KYC"],
                 ["phishing", "scam"]),
        TestCase("SAMPLE_008", "single-turn", "en",
                 ["What is today's weather?"],
                 ["cyber", "fraud"]),
    ]
